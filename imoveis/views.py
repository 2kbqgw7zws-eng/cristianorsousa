from django.shortcuts import render
from django.http import HttpResponse
from django.db.models import Sum, Count, Avg, Max
from django.db.models.functions import TruncMonth
from .models import Imovel, Locacao
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill
from django.template.loader import get_template
from xhtml2pdf import pisa

def buscar_dados_relatorio(ano_selecionado):
    # 1. Definir o período de análise (YTD - Year to Date)
    hoje = datetime.date.today()
    inicio_ano = datetime.date(ano_selecionado, 1, 1)
    
    if ano_selecionado == hoje.year:
        fim_periodo = hoje # Se é o ano atual, conta até hoje
    elif ano_selecionado < hoje.year:
        fim_periodo = datetime.date(ano_selecionado, 12, 31) # Ano passado conta o ano todo
    else:
        fim_periodo = inicio_ano # Futuro (ainda não começou a contar dias desocupados)

    # Total de dias corridos no período considerado
    dias_corridos_ano = (fim_periodo - inicio_ano).days + 1
    if dias_corridos_ano < 0: dias_corridos_ano = 0

    # 2. Buscar locações (filtrando pela data de entrada)
    locacoes_ano = Locacao.objects.filter(data_entrada__year=ano_selecionado)

    # 3. Processar dados por imóvel
    dados_imoveis = []
    imoveis = Imovel.objects.all()

    # Acumuladores Gerais
    total_fat_geral = 0
    total_dias_ocupados_geral = 0
    total_locacoes_geral = 0
    soma_ticket_medio = 0
    imoveis_com_locacao = 0

    for imovel in imoveis:
        locs_imovel = locacoes_ano.filter(imovel=imovel)
        
        # Calcular dias ocupados reais (Data Saída - Data Entrada)
        dias_ocupados = 0
        faturamento = 0
        qtd_locacoes = locs_imovel.count()

        for loc in locs_imovel:
            if loc.data_saida and loc.data_entrada:
                dias = (loc.data_saida - loc.data_entrada).days
                if dias > 0:
                    dias_ocupados += dias
            faturamento += loc.valor_cobrado_diaria or 0 # Assumindo que este campo é o valor TOTAL da reserva ou diária? 
            # *Nota: Se 'valor_cobrado_diaria' for só a diária, precisamos multiplicar pelos dias.
            # Vou assumir que sua lógica anterior (Sum) estava correta e esse campo já é o valor a somar.
        
        # Dias Desocupados (Vacância)
        dias_desocupados = dias_corridos_ano - dias_ocupados
        if dias_desocupados < 0: dias_desocupados = 0 # Segurança

        # Agrupamento Mensal
        meses = locs_imovel.annotate(mes_ref=TruncMonth('data_entrada')).values('mes_ref').annotate(
            qtd=Count('id'), total=Sum('valor_cobrado_diaria'),
            media=Avg('valor_cobrado_diaria'), maior=Max('valor_cobrado_diaria')
        ).order_by('mes_ref')

        dados_imoveis.append({
            'nome': imovel.nome,
            'faturamento_total': faturamento,
            'dias_ocupados': dias_ocupados,
            'dias_desocupados': dias_desocupados,
            'total_locacoes': qtd_locacoes,
            'ticket_medio': (faturamento / qtd_locacoes) if qtd_locacoes > 0 else 0,
            'meses': meses
        })

        # Somar aos totais gerais
        total_fat_geral += faturamento
        total_dias_ocupados_geral += dias_ocupados
        total_locacoes_geral += qtd_locacoes
        if qtd_locacoes > 0:
            soma_ticket_medio += (faturamento / qtd_locacoes)
            imoveis_com_locacao += 1

    # Cálculos Finais Gerais
    ticket_medio_geral = (soma_ticket_medio / imoveis_com_locacao) if imoveis_com_locacao > 0 else 0
    
    # "Dias Desocupados Geral" é a soma de dias ociosos de TODO o portfólio
    # Ex: Se tenho 2 imóveis e passou 1 dia. Se nenhum alugou, tenho 2 dias desocupados totais.
    total_dias_potenciais = dias_corridos_ano * imoveis.count()
    total_dias_desocupados_geral = total_dias_potenciais - total_dias_ocupados_geral
    if total_dias_desocupados_geral < 0: total_dias_desocupados_geral = 0

    geral = {
        'total_faturado': total_fat_geral,
        'total_locacoes': total_locacoes_geral,
        'ticket_medio': ticket_medio_geral,
        'dias_ocupados': total_dias_ocupados_geral,
        'dias_desocupados': total_dias_desocupados_geral
    }
    
    return {
        'ano': ano_selecionado,
        'ano_anterior': ano_selecionado - 1,
        'ano_proximo': ano_selecionado + 1,
        'geral': geral,
        'dados_imoveis': dados_imoveis,
        'periodo_dias': dias_corridos_ano
    }

# --- VIEW 1: Tela ---
def relatorio_geral(request):
    hoje = datetime.date.today()
    try:
        ano = int(request.GET.get('ano', hoje.year))
    except ValueError:
        ano = hoje.year
    contexto = buscar_dados_relatorio(ano)
    return render(request, 'relatorio.html', contexto)

# --- VIEW 2: PDF ---
def download_relatorio_pdf(request):
    hoje = datetime.date.today()
    try:
        ano = int(request.GET.get('ano', hoje.year))
    except ValueError:
        ano = hoje.year
    contexto = buscar_dados_relatorio(ano)
    template_path = 'relatorio_pdf.html'
    template = get_template(template_path)
    html = template.render(contexto)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Relatorio_{ano}.pdf"'
    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err: return HttpResponse('Erros ao gerar PDF')
    return response

# --- VIEW 3: EXCEL ---
def download_relatorio_excel(request):
    hoje = datetime.date.today()
    try:
        ano = int(request.GET.get('ano', hoje.year))
    except ValueError:
        ano = hoje.year

    dados = buscar_dados_relatorio(ano)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Relatorio {ano}"

    bold_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="2c3e50")
    header_font = Font(bold=True, color="FFFFFF")
    money_format = 'R$ #,##0.00'

    ws['A1'] = f"RELATÓRIO DE PERFORMANCE - {ano}"
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:E1')

    # Cabeçalho Geral
    headers_geral = ["Faturamento Total", "Total Locações", "Dias Ocupados", "Dias Desocupados", "Ticket Médio"]
    for i, h in enumerate(headers_geral, 1):
        cell = ws.cell(row=3, column=i, value=h)
        cell.font = bold_font
    
    ws['A4'] = dados['geral']['total_faturado']; ws['A4'].number_format = money_format
    ws['B4'] = dados['geral']['total_locacoes']
    ws['C4'] = dados['geral']['dias_ocupados']
    ws['D4'] = dados['geral']['dias_desocupados']
    ws['E4'] = dados['geral']['ticket_medio']; ws['E4'].number_format = money_format

    # Imóveis
    row = 7
    for imovel in dados['dados_imoveis']:
        ws.cell(row=row, column=1, value=imovel['nome']).font = Font(size=12, bold=True)
        # Resumo do Imóvel na linha
        ws.cell(row=row, column=2, value=f"Fat: R$ {imovel['faturamento_total']}")
        ws.cell(row=row, column=3, value=f"Ocup: {imovel['dias_ocupados']} dias")
        ws.cell(row=row, column=4, value=f"Livres: {imovel['dias_desocupados']} dias")
        row += 1

        headers = ["Mês", "Reservas", "Média Diária", "Maior Diária"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
        row += 1

        if not imovel['meses']:
            ws.cell(row=row, column=1, value="Sem dados.")
            row += 1
        else:
            for mes in imovel['meses']:
                ws.cell(row=row, column=1, value=mes['mes_ref'].strftime("%B/%Y"))
                ws.cell(row=row, column=2, value=mes['qtd'])
                ws.cell(row=row, column=3, value=mes['media']).number_format = money_format
                ws.cell(row=row, column=4, value=mes['maior']).number_format = money_format
                row += 1
        row += 2

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Relatorio_Completo_{ano}.xlsx"'
    wb.save(response)
    return response